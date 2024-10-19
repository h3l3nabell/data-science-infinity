# -*- coding: utf-8 -*-
"""
Spyder Editor

Cleaning Up LinkedIn Search Results copypasted into excel
"""


import openpyxl
import pandas as pd

wb = openpyxl.load_workbook('Data Scientists.xlsx')
sheets = wb.sheetnames
ws = wb[sheets[0]]
print(ws.cell(row=1, column=1).hyperlink.target)

doc_row = 1
doc_col = 1
last_row = 602

end_of_record_values = ["Connect", "Follow"]
new_doc_df = pd.DataFrame({'Name': [], 'Hyperlink': [], 'Current': []})

while doc_row <= last_row:
    cell = ws.cell(row=doc_row, column=doc_col)
    while not cell.value.contains(end_of_record_values):
        if cell.hyperlink != None:
            link = cell.hyperlink.target
